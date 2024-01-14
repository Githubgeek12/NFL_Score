import asyncio
import functools
import time
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from pyppeteer import launch
from bs4 import BeautifulSoup
from pyppeteer_stealth import stealth
import requests
import NFL_2
from io import StringIO


path = '/Users/trishika/PycharmProjects/pythonProject'
place = 'football'
f_path = '/Users/trishika/PycharmProjects/pythonProject/data/output_' + place + '.xlsx'


def retry(max_retries, delay=1):
    def decorator(func):
        @functools.wraps(func)
        async def wrapper(*args, **kwargs):
            for _ in range(max_retries):
                try:
                    result = await func(*args, **kwargs)
                    return result
                except Exception as e:
                    print(f"Retrying after error: {e}")
                    time.sleep(delay)
            raise Exception(f"Max retries reached for {func.__name__}")

        return wrapper

    return decorator


async def main():
    await scrape_data()


@retry(max_retries=3)
async def scrape_data():
    try:
        browser = await launch({'headless': True})
        page = await browser.newPage()
    except Exception:
        raise Exception('Error launching browser')
    await stealth(page)
    site = 'https://www.footballdb.com/games/index.html'
    try:
        await page.goto(site, {'waitUntil': 'domcontentloaded', 'timeout': 30000})
        await page.setViewport({"width": 1800, "height": 900})
        html = await page.content()
    except Exception:
        raise Exception('error getting page content')
    soup = BeautifulSoup(html, "html.parser")
    table = soup.find_all('table', class_='statistics')
    wb = Workbook()
    n = 1
    for idx, t in enumerate(table):
        strio = StringIO(str(t))
        df = pd.read_html(strio)[0]

        # Create a new Excel workbook and add a worksheet
        week = f'week{n}'
        w = wb.create_sheet(title=week)
        wb.active = w
        n += 1

        # Add the DataFrame to the worksheet
        for row in dataframe_to_rows(df, index=False, header=True):
            w.append(row)

        link = t.find_all('a')
        list_1 = []
        list_2 = []
        x = 0+(3*idx)
        for index, _ in enumerate(link, start=1):
            cell_coord = 'G'+str(index+1)
            link_1 = site[:-len('/games/index.html')]+_['href']
            w[cell_coord] = link_1
            dt = t.select_one('#leftcol > table:nth-child('+str(8+x)+') > tbody > tr:nth-child('+str(index)+') > td:nth-child(1) > span.hidden-xs').get_text()
            list_1.append(link_1)
            list_2.append(dt)

        wb1 = Workbook()
        tasks = [get_summary(site_1=_, name=week, _date=__, wb=wb1) for _, __ in zip(list_1, list_2)]
        await asyncio.gather(*tasks)
        print(f'{week} data download')

        # Save the workbook to an Excel file
        wb.save(f_path)

    await asyncio.sleep(5)


async def get_summary(site_1, name, _date, wb):
    await NFL_2.demo3(site_1=site_1, name=name, _date=_date, wb=wb)


if __name__ == '__main__':
    asyncio.get_event_loop().run_until_complete(main())
